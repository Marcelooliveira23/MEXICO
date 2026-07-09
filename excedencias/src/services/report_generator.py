"""
Gerador de Relatórios
Exporta resultados de análises para PDF, Excel e TXT
"""
from datetime import datetime
from pathlib import Path
from typing import List, Optional
import pandas as pd
from utils.logger import logger


class ReportGenerator:
    """Gera relatórios de análise em diferentes formatos"""
    
    def __init__(self):
        """Inicializa gerador de relatórios"""
        self.output_dir = Path("reports")
        self.output_dir.mkdir(exist_ok=True)
        logger.info("ReportGenerator inicializado")
    
    def generate_excel_report(
        self,
        data: pd.DataFrame,
        analysis_results: List,
        aircraft_family: str,
        event_type: str,
        output_path: Optional[str] = None
    ) -> str:
        """
        Gera relatório Excel com dados e análise
        
        Args:
            data: DataFrame com dados de voo
            analysis_results: Lista de resultados de análise
            aircraft_family: Família da aeronave
            event_type: Tipo de evento analisado
            output_path: Caminho do arquivo de saída (opcional)
            
        Returns:
            Caminho do arquivo gerado
        """
        try:
            # Importa openpyxl apenas quando necessário
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Define caminho de saída
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"analise_{aircraft_family}_{event_type}_{timestamp}.xlsx"
                output_path = self.output_dir / filename
            
            # Cria workbook
            wb = Workbook()
            
            # ABA 1: Resumo da Análise
            ws_summary = wb.active
            ws_summary.title = "Resumo"
            
            # Cabeçalho
            ws_summary['A1'] = "RELATÓRIO DE ANÁLISE DE INSPEÇÃO"
            ws_summary['A1'].font = Font(size=16, bold=True)
            ws_summary['A1'].alignment = Alignment(horizontal='center')
            ws_summary.merge_cells('A1:E1')
            
            # Informações gerais
            row = 3
            info = [
                ('Aeronave:', aircraft_family),
                ('Evento:', event_type),
                ('Data:', datetime.now().strftime("%d/%m/%Y %H:%M")),
                ('Total de Registros:', len(data)),
                ('Resultados Encontrados:', len(analysis_results))
            ]
            
            for label, value in info:
                ws_summary[f'A{row}'] = label
                ws_summary[f'A{row}'].font = Font(bold=True)
                ws_summary[f'B{row}'] = value
                row += 1
            
            # Tabela de resultados
            row += 2
            ws_summary[f'A{row}'] = "RESULTADOS DA ANÁLISE"
            ws_summary[f'A{row}'].font = Font(size=14, bold=True)
            row += 1
            
            # Cabeçalhos da tabela
            headers = ['Severidade', 'Descrição', 'Valor', 'Limite', 'Recomendação']
            for col, header in enumerate(headers, start=1):
                cell = ws_summary.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            row += 1
            
            # Dados dos resultados
            for result in analysis_results:
                severity_colors = {
                    'CRITICO': 'FF0000',
                    'ALERTA': 'FFA500',
                    'INFO': '008000'
                }
                
                ws_summary.cell(row=row, column=1, value=result.severity)
                ws_summary.cell(row=row, column=2, value=result.message)
                ws_summary.cell(row=row, column=3, value=str(result.value))
                ws_summary.cell(row=row, column=4, value=str(result.limit))
                ws_summary.cell(row=row, column=5, value=result.recommendation)
                
                # Cor de fundo baseada na severidade
                fill_color = severity_colors.get(result.severity, 'FFFFFF')
                for col in range(1, 6):
                    ws_summary.cell(row=row, column=col).fill = PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type="solid"
                    )
                    if result.severity in ['CRITICO', 'ALERTA']:
                        ws_summary.cell(row=row, column=col).font = Font(color="FFFFFF", bold=True)
                
                row += 1
            
            # Ajusta largura das colunas
            ws_summary.column_dimensions['A'].width = 15
            ws_summary.column_dimensions['B'].width = 40
            ws_summary.column_dimensions['C'].width = 15
            ws_summary.column_dimensions['D'].width = 15
            ws_summary.column_dimensions['E'].width = 50
            
            # ABA 2: Dados Completos
            ws_data = wb.create_sheet(title="Dados Completos")
            
            # Adiciona DataFrame
            for r_idx, row_data in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                for c_idx, value in enumerate(row_data, 1):
                    cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Formata cabeçalho
                    if r_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center')
            
            # Ajusta largura das colunas
            for column in ws_data.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_data.column_dimensions[column_letter].width = adjusted_width
            
            # Salva arquivo
            wb.save(output_path)
            logger.success(f"Relatório Excel gerado: {output_path}")
            
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório Excel: {e}")
            raise
    
    def generate_txt_report(
        self,
        data: pd.DataFrame,
        analysis_results: List,
        aircraft_family: str,
        event_type: str,
        output_path: Optional[str] = None
    ) -> str:
        """
        Gera relatório em formato TXT
        
        Args:
            data: DataFrame com dados de voo
            analysis_results: Lista de resultados de análise
            aircraft_family: Família da aeronave
            event_type: Tipo de evento analisado
            output_path: Caminho do arquivo de saída (opcional)
            
        Returns:
            Caminho do arquivo gerado
        """
        try:
            # Define caminho de saída
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"analise_{aircraft_family}_{event_type}_{timestamp}.txt"
                output_path = self.output_dir / filename
            
            with open(output_path, 'w', encoding='utf-8') as f:
                # Cabeçalho
                f.write("=" * 80 + "\n")
                f.write("RELATÓRIO DE ANÁLISE DE INSPEÇÃO DE AERONAVES\n")
                f.write("=" * 80 + "\n\n")
                
                # Informações gerais
                f.write(f"Aeronave: {aircraft_family}\n")
                f.write(f"Evento: {event_type}\n")
                f.write(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
                f.write(f"Total de Registros: {len(data)}\n")
                f.write(f"Resultados Encontrados: {len(analysis_results)}\n\n")
                
                # Resultados da análise
                f.write("-" * 80 + "\n")
                f.write("RESULTADOS DA ANÁLISE\n")
                f.write("-" * 80 + "\n\n")
                
                for i, result in enumerate(analysis_results, 1):
                    f.write(f"{i}. [{result.severity}] {result.message}\n")
                    f.write(f"   Valor: {result.value}\n")
                    f.write(f"   Limite: {result.limit}\n")
                    f.write(f"   Recomendação: {result.recommendation}\n\n")
                
                # Estatísticas dos dados
                f.write("-" * 80 + "\n")
                f.write("ESTATÍSTICAS DOS DADOS\n")
                f.write("-" * 80 + "\n\n")
                
                # Colunas numéricas
                numeric_cols = data.select_dtypes(include=['float64', 'int64']).columns
                for col in numeric_cols:
                    f.write(f"{col}:\n")
                    f.write(f"  Média: {data[col].mean():.2f}\n")
                    f.write(f"  Mínimo: {data[col].min():.2f}\n")
                    f.write(f"  Máximo: {data[col].max():.2f}\n")
                    f.write(f"  Desvio Padrão: {data[col].std():.2f}\n\n")
                
                # Rodapé
                f.write("=" * 80 + "\n")
                f.write("FIM DO RELATÓRIO\n")
                f.write("=" * 80 + "\n")
            
            logger.success(f"Relatório TXT gerado: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório TXT: {e}")
            raise
    
    def generate_pdf_report(
        self,
        data: pd.DataFrame,
        analysis_results: List,
        aircraft_family: str,
        event_type: str,
        output_path: Optional[str] = None
    ) -> str:
        """
        Gera relatório em formato PDF
        
        Args:
            data: DataFrame com dados de voo
            analysis_results: Lista de resultados de análise
            aircraft_family: Família da aeronave
            event_type: Tipo de evento analisado
            output_path: Caminho do arquivo de saída (opcional)
            
        Returns:
            Caminho do arquivo gerado
        """
        try:
            # Importa reportlab apenas quando necessário
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            
            # Define caminho de saída
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"analise_{aircraft_family}_{event_type}_{timestamp}.pdf"
                output_path = self.output_dir / filename
            
            # Cria documento
            doc = SimpleDocTemplate(
                str(output_path),
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Elementos do documento
            elements = []
            styles = getSampleStyleSheet()
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#366092'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            title = Paragraph("RELATÓRIO DE ANÁLISE DE INSPEÇÃO", title_style)
            elements.append(title)
            
            # Informações gerais
            info_data = [
                ['Aeronave:', aircraft_family],
                ['Evento:', event_type],
                ['Data:', datetime.now().strftime("%d/%m/%Y %H:%M")],
                ['Total de Registros:', str(len(data))],
                ['Resultados:', str(len(analysis_results))]
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('FONT', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONT', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#366092')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(info_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Resultados da análise
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#366092'),
                spaceAfter=12
            )
            subtitle = Paragraph("RESULTADOS DA ANÁLISE", subtitle_style)
            elements.append(subtitle)
            
            # Tabela de resultados
            results_data = [['Severidade', 'Descrição', 'Valor', 'Limite', 'Recomendação']]
            
            for result in analysis_results:
                results_data.append([
                    result.severity,
                    result.message[:50] + '...' if len(result.message) > 50 else result.message,
                    str(result.value),
                    str(result.limit),
                    result.recommendation[:60] + '...' if len(result.recommendation) > 60 else result.recommendation
                ])
            
            results_table = Table(results_data, colWidths=[1*inch, 1.8*inch, 1*inch, 1*inch, 1.8*inch])
            
            # Estilo da tabela
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
            
            # Cores de fundo por severidade
            for i, result in enumerate(analysis_results, 1):
                if result.severity == 'CRITICO':
                    table_style.append(('BACKGROUND', (0, i), (-1, i), colors.red))
                    table_style.append(('TEXTCOLOR', (0, i), (-1, i), colors.whitesmoke))
                elif result.severity == 'ALERTA':
                    table_style.append(('BACKGROUND', (0, i), (-1, i), colors.orange))
                elif result.severity == 'INFO':
                    table_style.append(('BACKGROUND', (0, i), (-1, i), colors.lightgreen))
            
            results_table.setStyle(TableStyle(table_style))
            elements.append(results_table)
            
            # Constrói PDF
            doc.build(elements)
            logger.success(f"Relatório PDF gerado: {output_path}")
            
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório PDF: {e}")
            raise
